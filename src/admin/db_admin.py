"""
Generic database idarəetməsi: cədvəllərin avtomatik aşkarlanması və
təhlükəsiz CRUD.

Təhlükəsizlik modeli:
  * Cədvəl və sütun adları HEÇ VAXT istifadəçidən birbaşa SQL-ə düşmür —
    əvvəlcə information_schema-dan oxunan real adlarla tutuşdurulur
    (whitelist), yalnız uyğun gələnlər istifadə olunur.
  * Dəyərlər həmişə parametrləşdirilmiş sorğu ilə ötürülür.
"""

import csv
import io
import json
from typing import Any

from db.connection import get_conn
from utils.logger import get_logger

logger = get_logger("AdminDB")

# Generic CRUD-dan qorunan sistem cədvəlləri:
#   admin_users — yalnız Users bölməsindən (parol hash-ləri korlanmasın)
#   audit_log   — audit jurnalı dəyişdirilə bilməz (təhlükəsizlik tələbi)
_PROTECTED_TABLES = {"admin_users", "audit_log"}


def list_tables() -> list[dict]:
    """Bütün istifadəçi cədvəllərini sətir sayı ilə qaytarır."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT c.relname AS name,
                   c.reltuples::bigint AS approx_rows
            FROM pg_class c
            JOIN pg_namespace n ON n.oid = c.relnamespace
            WHERE n.nspname = 'public' AND c.relkind = 'r'
            ORDER BY c.relname
        """).fetchall()
    return [dict(r) for r in rows]


def table_columns(table: str) -> list[dict]:
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT column_name AS name, data_type AS type,
                   is_nullable = 'YES' AS nullable,
                   column_default IS NOT NULL AS has_default
            FROM information_schema.columns
            WHERE table_schema = 'public' AND table_name = %s
            ORDER BY ordinal_position
        """, (table,)).fetchall()
    return [dict(r) for r in rows]


def _safe_table(table: str) -> str:
    names = {t["name"] for t in list_tables()}
    if table not in names:
        raise ValueError(f"Cədvəl tapılmadı: {table}")
    return table


def _safe_columns(table: str, cols: list[str]) -> list[str]:
    real = {c["name"] for c in table_columns(table)}
    return [c for c in cols if c in real]


def _pk_column(table: str) -> str:
    """Primary key sütununu tapır (yoxdursa 'id' fərz edilir)."""
    with get_conn() as conn:
        row = conn.execute("""
            SELECT a.attname AS name
            FROM pg_index i
            JOIN pg_attribute a ON a.attrelid = i.indrelid AND a.attnum = ANY(i.indkey)
            WHERE i.indrelid = %s::regclass AND i.indisprimary
            LIMIT 1
        """, (table,)).fetchone()
    return row["name"] if row else "id"


def query_rows(
    table: str, page: int = 1, per_page: int = 25,
    sort: str | None = None, order: str = "asc",
    search: str | None = None, filters: dict[str, str] | None = None,
) -> dict:
    """Pagination + sort + axtarış + sütun filtri ilə sətirlər."""
    table = _safe_table(table)
    cols = [c["name"] for c in table_columns(table)]
    where, params = [], []

    if search:
        text_cols = [c["name"] for c in table_columns(table)
                     if c["type"] in ("text", "character varying", "jsonb")]
        if text_cols:
            where.append("(" + " OR ".join(
                f'"{c}"::text ILIKE %s' for c in text_cols) + ")")
            params.extend([f"%{search}%"] * len(text_cols))

    for col, val in (filters or {}).items():
        if col in cols and val != "":
            where.append(f'"{col}"::text ILIKE %s')
            params.append(f"%{val}%")

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    sort_col = sort if sort in cols else _pk_column(table)
    order_sql = "DESC" if order.lower() == "desc" else "ASC"
    offset = (max(page, 1) - 1) * per_page

    with get_conn() as conn:
        total = conn.execute(
            f'SELECT count(*) AS n FROM "{table}" {where_sql}', params
        ).fetchone()["n"]
        rows = conn.execute(
            f'SELECT * FROM "{table}" {where_sql} '
            f'ORDER BY "{sort_col}" {order_sql} LIMIT %s OFFSET %s',
            params + [per_page, offset],
        ).fetchall()
    return {
        "rows": json.loads(json.dumps([dict(r) for r in rows], default=str, ensure_ascii=False)),
        "total": total, "page": page, "per_page": per_page, "columns": cols,
        "pk": _pk_column(table),
    }


def insert_row(table: str, data: dict[str, Any]) -> dict:
    table = _safe_table(table)
    if table in _PROTECTED_TABLES:
        raise ValueError("Bu cədvəl yalnız Users bölməsindən idarə olunur")
    cols = _safe_columns(table, list(data.keys()))
    if not cols:
        raise ValueError("Keçərli sütun verilməyib")
    vals = [data[c] if data[c] != "" else None for c in cols]
    col_sql = ", ".join(f'"{c}"' for c in cols)
    ph = ", ".join(["%s"] * len(cols))
    with get_conn() as conn:
        row = conn.execute(
            f'INSERT INTO "{table}" ({col_sql}) VALUES ({ph}) RETURNING *', vals
        ).fetchone()
    return json.loads(json.dumps(dict(row), default=str, ensure_ascii=False))


def update_row(table: str, pk_value: str, data: dict[str, Any]) -> dict:
    table = _safe_table(table)
    if table in _PROTECTED_TABLES:
        raise ValueError("Bu cədvəl yalnız Users bölməsindən idarə olunur")
    pk = _pk_column(table)
    cols = [c for c in _safe_columns(table, list(data.keys())) if c != pk]
    if not cols:
        raise ValueError("Keçərli sütun verilməyib")
    set_sql = ", ".join(f'"{c}" = %s' for c in cols)
    vals = [data[c] if data[c] != "" else None for c in cols]
    with get_conn() as conn:
        row = conn.execute(
            f'UPDATE "{table}" SET {set_sql} WHERE "{pk}"::text = %s RETURNING *',
            vals + [str(pk_value)],
        ).fetchone()
    if not row:
        raise ValueError("Sətir tapılmadı")
    return json.loads(json.dumps(dict(row), default=str, ensure_ascii=False))


def delete_rows(table: str, pk_values: list[str]) -> int:
    table = _safe_table(table)
    if table in _PROTECTED_TABLES:
        raise ValueError("Bu cədvəl yalnız Users bölməsindən idarə olunur")
    pk = _pk_column(table)
    with get_conn() as conn:
        cur = conn.execute(
            f'DELETE FROM "{table}" WHERE "{pk}"::text = ANY(%s)',
            ([str(v) for v in pk_values],),
        )
        return cur.rowcount


# ── Export / Import ────────────────────────────────────────────────────────

def export_rows(table: str, fmt: str) -> tuple[bytes, str, str]:
    """Bütün cədvəli csv/json/xlsx formatında qaytarır:
    (bytes, mime, filename)."""
    table = _safe_table(table)
    with get_conn() as conn:
        rows = [dict(r) for r in conn.execute(f'SELECT * FROM "{table}"').fetchall()]
    rows = json.loads(json.dumps(rows, default=str, ensure_ascii=False))

    if fmt == "json":
        data = json.dumps(rows, ensure_ascii=False, indent=2).encode("utf-8")
        return data, "application/json", f"{table}.json"

    cols = [c["name"] for c in table_columns(table)]
    if fmt == "csv":
        buf = io.StringIO()
        w = csv.DictWriter(buf, fieldnames=cols)
        w.writeheader()
        w.writerows(rows)
        return buf.getvalue().encode("utf-8-sig"), "text/csv", f"{table}.csv"

    if fmt == "xlsx":
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = table[:31]
        ws.append(cols)
        for r in rows:
            ws.append([r.get(c) for c in cols])
        buf = io.BytesIO()
        wb.save(buf)
        return (buf.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                f"{table}.xlsx")

    raise ValueError(f"Naməlum format: {fmt}")


def import_rows(table: str, filename: str, content: bytes) -> dict:
    """CSV/XLSX faylından sətirlər əlavə edir. Başlıq sətri sütun adları
    ilə üst-üstə düşməlidir; naməlum sütunlar ötürülür."""
    table = _safe_table(table)
    if table in _PROTECTED_TABLES:
        raise ValueError("Bu cədvəl yalnız Users bölməsindən idarə olunur")

    if filename.lower().endswith(".csv"):
        text = content.decode("utf-8-sig")
        records = list(csv.DictReader(io.StringIO(text)))
    elif filename.lower().endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = [str(h) for h in next(it)]
        records = [dict(zip(header, row)) for row in it]
    else:
        raise ValueError("Yalnız .csv və .xlsx dəstəklənir")

    inserted, errors = 0, []
    for i, rec in enumerate(records, 1):
        rec = {k: v for k, v in rec.items() if v not in (None, "")}
        try:
            insert_row(table, rec)
            inserted += 1
        except Exception as e:
            errors.append(f"Sətir {i}: {e}")
            if len(errors) >= 20:
                errors.append("...")
                break
    return {"inserted": inserted, "errors": errors}
